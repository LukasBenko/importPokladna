from __future__ import annotations

from typing import Dict, List, Optional, Tuple, Union
import xml.etree.ElementTree as ET

import openpyxl

from model import Doc, Item
from utils import fmt_date, indent, norm_ppd, to_num_str, to_str


SheetSel = Union[int, str]


def convert_xlsx_to_docs(xlsx_path: str, sheet: SheetSel = 0) -> List[Doc]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[int(sheet)]

    # find header row by "Skratka pokladne"
    header_row_idx = None
    header_map: Dict[str, int] = {}

    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(isinstance(v, str) and v.strip() == "Skratka pokladne" for v in row_vals):
            header_row_idx = r
            for c, v in enumerate(row_vals, start=1):
                if isinstance(v, str) and v.strip():
                    header_map[v.strip()] = c
            break

    if header_row_idx is None:
        raise ValueError('Nenašiel som riadok s hlavičkou "Skratka pokladne".')

    required = [
        "Skratka pokladne",
        "Druh PD",
        "Dátum PD",
        "Účel PD",
        "Skratka typu PPD",
        "Názov",
        "Suma",
        "OŠ",
    ]
    missing = [x for x in required if x not in header_map]
    if missing:
        raise ValueError(f"Chýbajú stĺpce: {missing}")

    col_komu_od = header_map.get("Komu/od") or header_map.get("Komu od") or header_map.get("Komu_od")
    col_eo = header_map.get("EO")

    docs: List[Doc] = []
    current: Optional[Doc] = None

    def key(d: Doc) -> Tuple[str, str, str, str, str]:
        return (d.skratka_pk, d.druh_pd, d.datum_pd, d.ucel_pd, d.komu_od)

    for r in range(header_row_idx + 1, ws.max_row + 1):
        skratka_pk = to_str(ws.cell(r, header_map["Skratka pokladne"]).value)
        druh_pd = to_str(ws.cell(r, header_map["Druh PD"]).value)
        datum_pd = fmt_date(ws.cell(r, header_map["Dátum PD"]).value)
        ucel_pd = to_str(ws.cell(r, header_map["Účel PD"]).value)
        komu_od = to_str(ws.cell(r, col_komu_od).value) if col_komu_od else ""

        skratka_typu_ppd = norm_ppd(ws.cell(r, header_map["Skratka typu PPD"]).value)
        nazov = to_str(ws.cell(r, header_map["Názov"]).value)
        suma = to_num_str(ws.cell(r, header_map["Suma"]).value)
        os_ = to_str(ws.cell(r, header_map["OŠ"]).value)
        eo_ = to_str(ws.cell(r, col_eo).value) if col_eo else ""

        if not any([skratka_pk, druh_pd, datum_pd, ucel_pd, komu_od, skratka_typu_ppd, nazov, suma, os_, eo_]):
            continue

        has_item = any([skratka_typu_ppd, nazov, suma, os_, eo_])
        has_header = any([skratka_pk, druh_pd, datum_pd, ucel_pd, komu_od])

        if has_header:
            d = Doc(
                skratka_pk=skratka_pk,
                druh_pd=druh_pd,
                datum_pd=datum_pd,
                ucel_pd=ucel_pd,
                komu_od=komu_od,
            )
            if current is None or key(d) != key(current):
                docs.append(d)
                current = d
        else:
            if current is None:
                raise ValueError(f"Riadok {r}: položka bez hlavičky (chýba doklad).")

        if has_item:
            current.items.append(
                Item(
                    skratka_typu_ppd=skratka_typu_ppd,
                    suma_ppd=suma,
                    poznamka_ppd=nazov,
                    skratka_os=os_,
                    skratka_eo=eo_,
                )
            )

    return docs


def docs_to_xml(
    docs: List[Doc],
    mandant_id: str = "1",
    encoding_decl: str = "windows-1250",
) -> bytes:
    root = ET.Element("pokladnicne_doklady")
    mandant = ET.SubElement(root, "mandant")
    mandant.set("mandant_id", str(mandant_id))

    for d in docs:
        doc_el = ET.SubElement(root, "pokladnicny_doklad")
        doc_el.set("skratka_pk", d.skratka_pk)
        doc_el.set("druh_pd", d.druh_pd)
        doc_el.set("datum_pd", d.datum_pd)
        doc_el.set("ucel_pd", d.ucel_pd)
        doc_el.set("komu_od", d.komu_od)

        for it in d.items:
            item_el = ET.SubElement(doc_el, "polozka_pd")
            if it.skratka_typu_ppd:
                item_el.set("skratka_typu_ppd", it.skratka_typu_ppd)
            if it.suma_ppd:
                item_el.set("suma_ppd", it.suma_ppd)
            if it.poznamka_ppd:
                item_el.set("poznamka_ppd", it.poznamka_ppd)
            if it.skratka_os:
                item_el.set("skratka_os", it.skratka_os)
            if it.skratka_eo:
                item_el.set("skratka_eo", it.skratka_eo)

            item_el.text = ""
            item_el.tail = "\n"

    indent(root)

    xml_unicode = ET.tostring(root, encoding="unicode")
    decl = f'<?xml version="1.0" encoding="{encoding_decl}" standalone="yes"?>\n'

    # finálny výstup cp1250
    out = decl.encode("cp1250", errors="xmlcharrefreplace") + xml_unicode.encode(
        "cp1250", errors="xmlcharrefreplace"
    )
    return out


def convert_xlsx_to_xml_file(
    xlsx_path: str,
    xml_path: str,
    sheet: SheetSel = 0,
    mandant_id: str = "1",
) -> Tuple[int, int]:
    docs = convert_xlsx_to_docs(xlsx_path, sheet=sheet)
    xml_bytes = docs_to_xml(docs, mandant_id=mandant_id)

    with open(xml_path, "wb") as f:
        f.write(xml_bytes)

    item_count = sum(len(d.items) for d in docs)
    return len(docs), item_count
